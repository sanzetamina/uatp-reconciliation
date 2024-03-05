import pandas as pd
import glob
from datetime import datetime
import openpyxl
from openpyxl.styles import Font
from openpyxl.comments import Comment


# read all excel files in the directory as input, excluding those with name containing 'ouput'
def read_input_files():
    input_files = [f for f in glob.glob("*.xls*") if "output" not in f.lower()]
    appended_data = []

    for f in input_files:
        print(f"Reading...... input file: {f}")
        appended_data.append(pd.read_excel(f))

    df = pd.concat(appended_data)
    return df


# formatting by:
# filling-in existing empty cells with a 0 value
# formatting "TRANSACTION NUMBER" (ETKTs) as "083-0000000000"
def format_dataframe(df):
    df.fillna("NIL", inplace=True)
    df["TRANSACTION NUMBER"] = df["TRANSACTION NUMBER"].apply("{:0>13}".format)
    df["TRANSACTION NUMBER"] = df["TRANSACTION NUMBER"].str.replace(
        r"(?<=^.{3})", r"-", regex=True
    )
    return df


# pivot table for most of the reconciliation work by matching REFUNDS to SALES
# by PNRs then ETKTS rows, and sort by column Total ascending
def create_pivot_table(df):
    pivot_table = pd.pivot_table(
        df,
        index=["CUSTOMER REFERENCE", "TRANSACTION NUMBER"],
        columns="TRANSACTION TYPE",
        values="BILLING VALUE",
        fill_value=".",
        aggfunc=sum,
        margins=True,
        margins_name="Total",
    )
    pivot_table.drop("Total", axis=0, inplace=True)
    return pivot_table.reset_index()


# Run a 2nd round of re-processing, let's call it "pnr_grouped", this time taking as source the "dfOutstandingTrxs" dataframe
# since there are many instances of PNRs with more than one ticket, they end up not being considered as settled (many false positives)
# by doing this additional round, we hope we can reduce the output of outstanding trx by removingn the false positives which are actually settled (via exchanged tickets)
# create new pivot table having dfOutstandingTrxs as input
def create_grouped_pivot_table(df):
    pivot_table = pd.pivot_table(
        df,
        index=["CUSTOMER REFERENCE"],
        values="Total",
        fill_value=".",
        aggfunc=sum,
        margins=True,
        margins_name="Total",
    )
    return pivot_table.reset_index()


# Write output excel
def write_to_excel(
    df,
    df_pivot,
    df_settled_trxs,
    df_outstanding_trxs,
    df_settled_pnrs,
    df_outstanding_pnrs,
):
    # Get the current date and time
    now = datetime.now()

    # Format the date and time as a string "yyyymmdd-hhmm"
    timestamp_str = now.strftime("%Y%m%d-%H%M")

    # Use the formatted string in your filename
    filename = f"Output-{timestamp_str}.xlsx"

    print(f"Creating new output file: {filename}")

    with pd.ExcelWriter(filename) as writer:
        df.to_excel(writer, "UATP Source", index=False)
        df_pivot.to_excel(writer, "Pivot", index=False)
        df_settled_trxs.to_excel(writer, "Settled Trxs", index=False)
        df_outstanding_trxs.to_excel(writer, "Outstanding Trxs", index=False)
        df_settled_pnrs.to_excel(writer, "Settled PNRs", index=False)
        df_outstanding_pnrs.to_excel(
            writer, "Outstanding PNRs", index=False, float_format="%.2f"
        )

    # Load the workbook and iterate over sheets to rename the 'Total' column
    book = openpyxl.load_workbook(filename)
    for sheetname in book.sheetnames:
        sheet = book[sheetname]
        for i, cell in enumerate(sheet[1], start=1):  # header in second row
            if cell.value == "Total":
                cell.value = "AU$ Total"
                break  # exit the loop once the cell is found and modified
    book.save(filename)

    return filename


def set_freeze_panes_and_columns_width(filename):
    book = openpyxl.load_workbook(filename)

    for sheet_name in book.sheetnames:
        sheet = book[sheet_name]

        # Freeze the first row
        sheet.freeze_panes = "A3"

        # Iterate over the columns and set their width
        for column in sheet.columns:
            max_length = 0
            column = [
                cell for cell in column if cell.row > 1
            ]  # Skip the first row where title will be inserted, so the column doesn't get too wide
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = max_length + 2
            sheet.column_dimensions[
                openpyxl.utils.get_column_letter(column[0].column)
            ].width = adjusted_width

    book.save(filename)


def add_titles_to_sheets(filename):
    book = openpyxl.load_workbook(filename)
    titles = {
        "UATP Source": "UATP Source as per the Input file/s",
        "Pivot": "Pivot Table to work out the Reconciliation. Matching Sales and Refunds by a combination of PNR and Ticket",
        "Settled Trxs": "Settled Tickets. For which 'Sales - Refunds = 0'",
        "Outstanding Trxs": "Outstanding Tickets. For which Sales and Refunds don't tally. Sorted by Total, from smallest to largest",
        "Settled PNRs": "Settled PNRs. For which all associated Tickets' Sales and Refunds add to Zero",
        "Outstanding PNRs": "Outstanding PNRs. For which associated tickets have outstanding amounts not tallied. Sorted by Total, from smallest to largest. 'NIL' groups all tickets with no know PNR",
    }

    for sheet_name, title in titles.items():
        sheet = book[sheet_name]

        # Insert a new row at the first position
        sheet.insert_rows(idx=1, amount=1)

        # Adding the title in the first cell of the new row
        title_cell = sheet.cell(row=1, column=1, value=title)

        # Setting the font style of the title to bold, italic, and larger size.
        title_cell.font = Font(bold=True, italic=True, size=14)

    book.save(filename)


def add_note_to_outstanding_pnrs_sheet(filename):
    book = openpyxl.load_workbook(filename)

    sheet = book["Outstanding PNRs"]

    # Adding a note to be placed next to the total row at the bottom of this worksheet
    row = sheet.max_row
    column = 3  # note needs to be placed in column C

    # Create a comment
    note = "<-- This Total should be the same as the sum of 'BILLING VALUE' column in 'UATP Source' to confirm everything went well. Please check!"

    # Place the note in the cell
    sheet.cell(row=row, column=column, value=note).font = Font(bold=True)

    book.save(filename)


def main():
    # to be able to see the full cols/row of dataframes on the terminal
    pd.set_option("display.max_rows", None)

    # Input
    try:
        df = read_input_files()
    except Exception as e:
        print(f"Error reading input files: {e}")
        return

    # Operation
    df = format_dataframe(df)

    # sometimes input files may have CR_NOTE and DB_NOTE as TRANSACTION TYPE, which we won't handle
    # instead, we'll just remove them from the dataframe before proceeding
    df = df[df["TRANSACTION TYPE"].isin(["SALES", "REFUND"])]

    df_pivot = create_pivot_table(df)
    print("--- Pivot Table ---")
    print(df_pivot)

    df_pivot.sort_values(by=["Total"], ascending=True, inplace=True)

    df_settled_trxs = df_pivot[df_pivot.Total == 0]
    df_outstanding_trxs = df_pivot[df_pivot.Total != 0]

    df_pnr_grouped = create_grouped_pivot_table(df_outstanding_trxs).round(2)
    df_pnr_grouped.sort_values(by="Total", ascending=True, inplace=True)

    df_settled_pnrs = df_pnr_grouped[df_pnr_grouped.Total == 0]
    df_outstanding_pnrs = df_pnr_grouped[df_pnr_grouped.Total != 0]
    df_outstanding_pnrs = df_outstanding_pnrs.sort_values(by="Total", ascending=True)

    # Output
    output_filename = write_to_excel(
        df,
        df_pivot,
        df_settled_trxs,
        df_outstanding_trxs,
        df_settled_pnrs,
        df_outstanding_pnrs,
    )

    # Output formatting
    print(f"Formatting output file...")
    add_titles_to_sheets(output_filename)
    set_freeze_panes_and_columns_width(output_filename)
    add_note_to_outstanding_pnrs_sheet(output_filename)
    print(f"Done! all finished.")
    # await for key stroke to exit
    input("Press any key to exit...")
    return


if __name__ == "__main__":
    main()
